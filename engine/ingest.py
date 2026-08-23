
from __future__ import annotations

import io
import re
from dataclasses import asdict, dataclass, field

import openpyxl

SHEET = "Course Distribution"
HEADER_ROW = 5
DATA_START = 6

SERVICE_SECTION = re.compile(r"^(GED|WD|WE)", re.IGNORECASE)


@dataclass
class Session:
    sid: int
    code: str
    title: str
    teacher: str
    batch: str
    section: str
    cohort: str
    is_lab: bool
    is_service: bool
    occurrence: int


def _norm_batch(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return str(int(v))
    s = str(v).strip()
    return s or None


def _last_digit(code: str) -> int | None:
    m = re.search(r"(\d+)\s*$", code or "")
    return int(m.group(1)[-1]) if m else None


def _header_map(ws) -> dict[str, int]:
    out: dict[str, int] = {}
    for c in range(1, (ws.max_column or 30) + 1):
        v = ws.cell(row=HEADER_ROW, column=c).value
        if v:
            out[str(v).strip().lower()] = c
    return out


def _col(hmap: dict[str, int], *prefixes: str) -> int | None:
    for p in prefixes:
        for k, v in hmap.items():
            if k.startswith(p.lower()):
                return v
    return None


def ingest_bytes(data: bytes) -> dict:
    return ingest_workbook(openpyxl.load_workbook(io.BytesIO(data), data_only=True))


def ingest_path(path: str) -> dict:
    return ingest_workbook(openpyxl.load_workbook(path, data_only=True))


def ingest_workbook(wb) -> dict:
    if SHEET not in wb.sheetnames:
        raise ValueError(f'Workbook has no "{SHEET}" sheet. Got: {wb.sheetnames}')
    ws = wb[SHEET]
    h = _header_map(ws)

    cB = _col(h, "batch")
    cS = _col(h, "section")
    cCode = _col(h, "course code")
    cTitle = _col(h, "course title")
    cDept = _col(h, "conducting")
    cTea = _col(h, "teacher")
    cDur = _col(h, "class duration")
    cWk = _col(h, "class/week", "class /week", "class per week")
    missing = [n for n, c in {"Batch": cB, "Section": cS, "Course Code": cCode,
                              "Teacher": cTea}.items() if c is None]
    if missing:
        raise ValueError(f"Could not find required column(s): {missing}. "
                         f"Headers seen: {sorted(h.keys())}")

    sessions: list[Session] = []
    excluded: list[dict] = []
    cohorts: set[str] = set()
    teachers_used: set[str] = set()
    warnings: list[str] = []
    sid = 0
    n_offerings = 0

    for r in range(DATA_START, (ws.max_row or DATA_START) + 1):
        code = ws.cell(row=r, column=cCode).value
        batch = _norm_batch(ws.cell(row=r, column=cB).value)
        if not code and not batch:
            continue
        code = str(code).strip() if code else None
        if not code:
            continue
        if not batch:
            excluded.append({"row": r, "code": code, "batch": None,
                             "section": None, "reason": "no_batch"})
            continue
        n_offerings += 1

        section = ws.cell(row=r, column=cS).value
        section = str(section).strip() if section is not None else ""
        title = ws.cell(row=r, column=cTitle).value if cTitle else None
        title = str(title).strip() if title else code
        teacher = ws.cell(row=r, column=cTea).value
        teacher = str(teacher).strip() if teacher is not None else ""

        if cWk:
            wk = ws.cell(row=r, column=cWk).value
            try:
                if wk is not None and int(float(wk)) != 2:
                    warnings.append(f"Row {r} ({code} {batch}-{section}): Class/Week={wk} (expected 2).")
            except (TypeError, ValueError):
                warnings.append(f"Row {r} ({code}): unreadable Class/Week '{wk}'.")
        if cDur:
            dur = ws.cell(row=r, column=cDur).value
            try:
                if dur is not None and abs(float(dur) - 1.5) > 1e-6:
                    warnings.append(f"Row {r} ({code}): Class Duration={dur} (expected 1.5).")
            except (TypeError, ValueError):
                pass

        if teacher == "":
            excluded.append({"row": r, "code": code, "batch": batch,
                             "section": section, "reason": "no_teacher"})
            continue

        ld = _last_digit(code)
        is_lab = ld is not None and ld % 2 == 0
        is_service = (not batch.isdigit()) or bool(SERVICE_SECTION.match(section))
        cohort = f"{batch}-{section}" if section else f"{batch}"

        teachers_used.add(teacher)
        if not is_service:
            cohorts.add(cohort)

        for occ in (1, 2):
            sessions.append(Session(
                sid=sid, code=code, title=title, teacher=teacher,
                batch=batch or "", section=section, cohort=cohort,
                is_lab=is_lab, is_service=is_service, occurrence=occ))
            sid += 1

    meta = {
        "offerings": n_offerings,
        "sessions": len(sessions),
        "render_sessions": sum(1 for s in sessions if not s.is_service),
        "service_sessions": sum(1 for s in sessions if s.is_service),
        "excluded": len(excluded),
        "cohorts": len(cohorts),
        "teachers": len(teachers_used),
    }
    return {
        "cohorts": _ordered_cohorts(cohorts),
        "sessions": [asdict(s) for s in sessions],
        "teachers_used": sorted(teachers_used),
        "excluded": excluded,
        "warnings": warnings,
        "meta": meta,
    }


def _ordered_cohorts(cohorts: set[str]) -> list[str]:
    """Canonical render order: batch descending (66 first), section ascending,
    with merged sections (B+C) and lettered sections sorted naturally."""
    def key(c: str):
        if "-" in c:
            b, s = c.split("-", 1)
        else:
            b, s = c, ""
        try:
            bnum = int(b)
        except ValueError:
            bnum = -1
        return (-bnum, s)
    return sorted(cohorts, key=key)


if __name__ == "__main__":
    import json
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else \
        "routine generation files/Main_Distribution_Summer25.xlsx"
    res = ingest_path(path)
    print(json.dumps(res["meta"], indent=2))
    print("cohorts:", res["cohorts"])
    print("excluded:", res["excluded"])
    print(f"warnings ({len(res['warnings'])}):")
    for w in res["warnings"][:15]:
        print("  -", w)
