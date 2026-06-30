
from __future__ import annotations

import io
from pathlib import Path

import openpyxl
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.styles import Border, Side

TEMPLATE = Path(__file__).parent / "templates" / "CSE_Routine_TEMPLATE.xlsx"

DAYS = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
HEADER_ROW = 2
DATA_FIRST_ROW = 3
DATA_LAST_ROW = 82
BREAK_COL = 7
SESSION_COLS = [4, 5, 6, 8, 9, 10, 11]
GRID_COLS = [2, 3] + SESSION_COLS
_THIN = Side(style="thin", color="FF000000")
_THICK = Side(style="thick", color="FF000000")


def _batch_cell(batch: str):
    """Write numeric batches as ints (66) not floats (66.0); keep text as-is."""
    try:
        return int(batch)
    except (TypeError, ValueError):
        return batch


def render_bytes(rows: list[dict], cohorts: list[str], config: dict,
                 template_path: Path | None = None) -> bytes:
    tpl = template_path or TEMPLATE
    if not Path(tpl).exists():
        raise FileNotFoundError(
            f"Template not found at {tpl}. Run tools/make_template.py once.")
    wb = openpyxl.load_workbook(tpl)

    pcol = {int(p["idx"]): openpyxl.utils.column_index_from_string(p["col"])
            for p in config["periods"]}

    row_of = {c: DATA_FIRST_ROW + i for i, c in enumerate(cohorts)}

    by_cohort: dict[str, list[dict]] = {}
    for r in rows:
        key = f"{r['batch']}-{r['section']}" if r["section"] else str(r["batch"])
        by_cohort.setdefault(key, []).append(r)

    for day in DAYS:
        if day not in wb.sheetnames:
            continue
        ws = wb[day]

        _write_period_headers(ws, day, config)

        for rr in range(DATA_FIRST_ROW, DATA_LAST_ROW + 1):
            for cc in SESSION_COLS:
                _safe_clear(ws, rr, cc)

        for rr in range(DATA_FIRST_ROW, DATA_LAST_ROW + 1):
            idx = rr - DATA_FIRST_ROW
            if idx < len(cohorts):
                c = cohorts[idx]
                batch, _, section = c.partition("-")
                ws.cell(row=rr, column=2, value=_batch_cell(batch))
                ws.cell(row=rr, column=3, value=section or None)
                nxt = cohorts[idx + 1] if idx + 1 < len(cohorts) else None
                last_of_batch = nxt is None or nxt.partition("-")[0] != batch
                _set_row_bottom(ws, rr, _THICK if last_of_batch else _THIN)
                ws.row_dimensions[rr].hidden = False
            else:
                _drop_data_row_merges(ws, rr)
                _safe_clear(ws, rr, 2)
                _safe_clear(ws, rr, 3)
                _set_row_bottom(ws, rr, None)
                ws.row_dimensions[rr].hidden = True

        for cohort, sess in by_cohort.items():
            rr = row_of.get(cohort)
            if rr is None:
                continue
            for s in sess:
                if s["day"] != day:
                    continue
                col = pcol.get(int(s["period"]))
                if col is None or col == BREAK_COL:
                    continue
                text = f"{s['subject_code']} {s['teacher_code']} {s['room']}"
                ws.cell(row=rr, column=col, value=text)

        last_data = min(DATA_FIRST_ROW + len(cohorts) - 1, DATA_LAST_ROW)
        if last_data >= DATA_FIRST_ROW:
            _rebuild_break_column(ws, day, last_data)

    _strip_legacy_scaffolding(wb)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _safe_clear(ws, row: int, col: int) -> None:
    """Set a cell to None unless it is a non-anchor cell of a merged range
    (writing those raises in openpyxl)."""
    from openpyxl.cell.cell import MergedCell
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = None


def _strip_legacy_scaffolding(wb) -> None:
    """Strip every macro-era interactive artifact the source workbook carried
    over, so the generated routine is a plain, static spreadsheet (the new
    system no longer uses any of this scaffolding):

      - AutoFilters + their auto-generated `_FilterDatabase` defined names,
      - conditional-formatting rules,
      - structured Tables (e.g. 'Teacher Day Off'),
      - broken `#REF!` defined names (dead dynamic ranges the old dropdowns fed
        from).

    None of these affect the routine's visible content, styling, or layout.
    They also make lightweight/mobile xlsx viewers blank the whole sheet, which
    is why the day-sheets came up empty on a phone while the artifact-free
    sheets rendered fine. (The template has no VBA — it is a plain .xlsx, not a
    macro-enabled .xlsm — so there is no executable code to remove.)"""
    for ws in wb.worksheets:
        ws.auto_filter.ref = None
        ws.conditional_formatting = ConditionalFormattingList()
        for table_name in list(ws.tables.keys()):
            del ws.tables[table_name]

    for name in [n for n, dn in wb.defined_names.items()
                 if dn.value and "#REF!" in dn.value]:
        del wb.defined_names[name]


def _fmt_clock(hhmm: str) -> tuple[str, str]:
    """'13:10' -> ('1:10', 'PM'). Hours <= 12 keep two digits (08:50, 12:35);
    afternoon hours convert to 12-hour without a leading zero (13->1), matching
    the source workbook's hand-typed header style."""
    h, m = map(int, hhmm.split(":"))
    ampm = "AM" if h < 12 else "PM"
    hh = f"{h:02d}" if h <= 12 else f"{h - 12}"
    return f"{hh}:{m:02d}", ampm


def _period_label(start: str, end: str) -> str:
    """Header label 'start-endAM/PM' (AM/PM taken from the end time)."""
    s, _ = _fmt_clock(start)
    e, ap = _fmt_clock(end)
    return f"{s}-{e}{ap}"


def _set_header_value(ws, row: int, col: int, text: str) -> None:
    """Set a header cell's text, preserving its style; skip merged interiors."""
    from openpyxl.cell.cell import MergedCell
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = text


def _write_period_headers(ws, day: str, config: dict) -> None:
    """Rewrite row 2's period-time labels from config (in-person periods +
    computed break). Online column and Batch/Section headers are left as-is."""
    in_person = sorted((p for p in config["periods"] if int(p["idx"]) <= 6),
                       key=lambda p: int(p["idx"]))
    drop_p4 = day == "Friday" and config.get("friday_no_p4", True)
    day_periods = [p for p in in_person if not (drop_p4 and int(p["idx"]) == 4)]
    for p in day_periods:
        col = openpyxl.utils.column_index_from_string(p["col"])
        _set_header_value(ws, HEADER_ROW, col, _period_label(p["start"], p["end"]))
    for a, b in zip(day_periods, day_periods[1:]):
        if a["end"] != b["start"]:
            _set_header_value(ws, HEADER_ROW, BREAK_COL,
                              _period_label(a["end"], b["start"]))
            break


def _rebuild_break_column(ws, day: str, last_data: int) -> None:
    """Re-merge the vertical BREAK column to span only the used data rows
    (3..last_data), so its 'BREAK' letters always fit the visible table no
    matter the cohort count. The template ships with the column pre-merged for
    the full capacity; this shrinks it to the live size. Style is copied from
    the template's existing break cell."""
    from copy import copy
    from openpyxl.cell.cell import MergedCell
    src = ws.cell(DATA_FIRST_ROW, BREAK_COL)
    font, align, fill, border = (copy(src.font), copy(src.alignment),
                                 copy(src.fill), copy(src.border))
    for m in list(ws.merged_cells.ranges):
        if m.min_col == BREAK_COL and DATA_FIRST_ROW <= m.min_row <= DATA_LAST_ROW:
            ws.unmerge_cells(str(m))
    for r in range(DATA_FIRST_ROW, DATA_LAST_ROW + 1):
        cell = ws.cell(r, BREAK_COL)
        if not isinstance(cell, MergedCell):
            cell.value = None

    def _style(cell, text):
        cell.font, cell.alignment = copy(font), copy(align)
        cell.fill, cell.border = copy(fill), copy(border)
        cell.value = text

    if day == "Friday":
        ws.merge_cells(start_row=DATA_FIRST_ROW, start_column=BREAK_COL,
                       end_row=last_data, end_column=BREAK_COL + 1)
        _style(ws.cell(DATA_FIRST_ROW, BREAK_COL), "BREAK")
        return
    letters = "BREAK"
    size = max(1, (last_data - DATA_FIRST_ROW + 1) // len(letters))
    start = DATA_FIRST_ROW
    for i, ch in enumerate(letters):
        end = last_data if i == len(letters) - 1 else min(last_data, start + size - 1)
        ws.merge_cells(start_row=start, start_column=BREAK_COL,
                       end_row=end, end_column=BREAK_COL)
        _style(ws.cell(start, BREAK_COL), ch)
        start = end + 1
        if start > last_data:
            break


def _drop_data_row_merges(ws, row: int) -> None:
    """Unmerge any leftover single-row merge inside the data columns of an
    empty trailing row (e.g. a stale 'A+B' B:C merge), so it can be fully
    blanked. Multi-row merges (the vertical BREAK block) are left intact."""
    for mr in list(ws.merged_cells.ranges):
        if (mr.min_row == row == mr.max_row
                and mr.min_col >= 2 and mr.max_col <= max(GRID_COLS)):
            ws.unmerge_cells(str(mr))


def _set_row_bottom(ws, row: int, side: Side | None) -> None:
    """Set the bottom border across the table's grid columns for one row,
    preserving each cell's existing left/right/top sides. `side=None` clears
    the bottom (used to wipe stray separators on empty trailing rows). Skips
    merged-range interior cells, which can't take an individual border."""
    from openpyxl.cell.cell import MergedCell
    for col in GRID_COLS:
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            continue
        b = cell.border
        cell.border = Border(left=b.left, right=b.right, top=b.top, bottom=side)


if __name__ == "__main__":
    import ingest
    import solver

    ds = ingest.ingest_path("routine generation files/Main_Distribution_Summer25.xlsx")
    cfg = solver.load_config()
    res = solver.solve(ds, cfg, time_limit_s=20.0)
    data = render_bytes(res["rows"], ds["cohorts"], cfg)
    out = Path(__file__).parent / "out_routine.xlsx"
    out.write_bytes(data)
    print(f"rendered {len(res['rows'])} sessions -> {out} ({len(data)} bytes)")
