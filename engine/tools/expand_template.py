
from __future__ import annotations

import sys
from copy import copy
from pathlib import Path

import openpyxl

TEMPLATE = Path(__file__).resolve().parent.parent / "templates" / "CSE_Routine_TEMPLATE.xlsx"
DAY_SHEETS = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

HEADER_ROW = 2
DATA_FIRST = 3
OLD_LAST = 59
FOOTER_START = 60
STYLE_COLS = [2, 3, 4, 5, 6, 8, 9, 10, 11]
BREAK_COL = 7
TARGET_ROWS = 80


def _copy_style(src, dst) -> None:
    if src.has_style:
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def _expand_sheet(ws, add: int, new_last: int) -> None:
    is_friday = ws.title == "Friday"
    ref = {c: ws.cell(DATA_FIRST, c) for c in STYLE_COLS}
    row_h = ws.row_dimensions[DATA_FIRST].height
    brk_src = ws.cell(DATA_FIRST, BREAK_COL)

    merges = [str(m) for m in ws.merged_cells.ranges]
    ws.insert_rows(FOOTER_START, add)

    for ref_str in merges:
        mr = openpyxl.worksheet.cell_range.CellRange(ref_str)
        keep = ((mr.min_row == 1) or
                (mr.min_row == HEADER_ROW and mr.min_col >= BREAK_COL))
        a_col = (mr.min_col == 1 == mr.max_col and mr.min_row == HEADER_ROW)
        is_break = (mr.min_col == BREAK_COL and mr.min_row >= DATA_FIRST)
        footer = (mr.min_row >= FOOTER_START)
        if keep:
            continue
        try:
            ws.unmerge_cells(ref_str)
        except (KeyError, ValueError):
            pass
        if a_col or is_break:
            pass
        elif footer:
            ws.merge_cells(start_row=mr.min_row + add, start_column=mr.min_col,
                           end_row=mr.max_row + add, end_column=mr.max_col)

    new_bus_last = OLD_LAST + add + 2
    ws.merge_cells(start_row=HEADER_ROW, start_column=1, end_row=new_bus_last, end_column=1)

    for r in range(FOOTER_START, new_last + 1):
        for c in STYLE_COLS:
            _copy_style(ref[c], ws.cell(r, c))
        if row_h:
            ws.row_dimensions[r].height = row_h
    for r in range(DATA_FIRST, new_last + 1):
        ws.row_dimensions[r].hidden = False

    if is_friday:
        ws.merge_cells(start_row=DATA_FIRST, start_column=BREAK_COL,
                       end_row=new_last, end_column=BREAK_COL + 1)
        a = ws.cell(DATA_FIRST, BREAK_COL); _copy_style(brk_src, a); a.value = "BREAK"
    else:
        span = new_last - DATA_FIRST + 1
        letters = ["B", "R", "E", "A", "K"]
        size = span // len(letters)
        start = DATA_FIRST
        for i, ch in enumerate(letters):
            end = new_last if i == len(letters) - 1 else start + size - 1
            ws.merge_cells(start_row=start, start_column=BREAK_COL,
                           end_row=end, end_column=BREAK_COL)
            a = ws.cell(start, BREAK_COL); _copy_style(brk_src, a); a.value = ch
            start = end + 1


def expand(out_path: Path, target_rows: int = TARGET_ROWS) -> None:
    wb = openpyxl.load_workbook(TEMPLATE)
    current = OLD_LAST - DATA_FIRST + 1
    if target_rows <= current:
        print(f"template already has {current} rows >= target {target_rows}; nothing to do.")
        return
    add = target_rows - current
    new_last = OLD_LAST + add
    for name in DAY_SHEETS:
        _expand_sheet(wb[name], add, new_last)
    wb.save(out_path)
    print(f"expanded {len(DAY_SHEETS)} day-sheets: {current} -> {target_rows} data rows "
          f"(rows {DATA_FIRST}..{new_last}); footer now starts at {new_last + 1}. "
          f"Saved -> {out_path}")


if __name__ == "__main__":
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else TEMPLATE
    tgt = int(sys.argv[2]) if len(sys.argv) > 2 else TARGET_ROWS
    expand(out, tgt)
