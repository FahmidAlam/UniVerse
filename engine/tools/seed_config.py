# ============================================================
# FILE: engine/tools/seed_config.py
# PURPOSE: One-off generator. Mines the two real workbooks for the
#   timetable engine's configuration and emits:
#     1) engine/config.json                    — engine-side fallback config
#     2) supabase/seed/seed_timetable_config.sql — DB seed for the
#        timetable_rooms / timetable_faculty / timetable_settings tables
#
#   Source of truth per element:
#     - teacher acronyms  -> Main_Distribution "Course Distribution" col Teacher
#     - teacher off-days  -> CSE_Routine "Teacher Day Off"  (TRUE = OFF)
#     - teacher names     -> CSE_Routine "Teacher Information" + "GED Teacher Info"
#     - rooms             -> CSE_Routine "Room Acronym" + mined from day-sheet cells
#     - periods / grid    -> CSE_Routine day-sheet headers (08:50 grid, Fri no P4)
#
# RUN (from engine/):  .venv/Scripts/python.exe tools/seed_config.py
# ============================================================

from __future__ import annotations

import json
import re
from collections import Counter
from pathlib import Path

import openpyxl

ENGINE_DIR = Path(__file__).resolve().parent.parent
REPO_DIR = ENGINE_DIR.parent
REF_DIR = ENGINE_DIR / "routine generation files"
DISTRIBUTION = REF_DIR / "Main_Distribution_Summer25.xlsx"
ROUTINE = REF_DIR / "CSE Routine Summer'25 Version 2.0.xlsx"

DAYS = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
PERIOD_COLS = [4, 5, 6, 8, 9, 10, 11]  # D,E,F,H,I,J,K (G=break)

# Authoritative room classification (spec §3.3, validated against mined cells).
LAB_ROOMS = {"ACL-1", "ACL-2", "ACL-3", "ACL-4", "NL", "GL", "ECL-1", "ECL-2"}
GALLERIES = {"G1", "G2", "G3"}

# Period grid read from the day-sheet headers (08:50 start). idx is 1-based
# period number; col is the Excel column letter in the output workbook.
PERIODS = [
    {"idx": 1, "label": "08:50-10:05", "start": "08:50", "end": "10:05", "col": "D"},
    {"idx": 2, "label": "10:05-11:20", "start": "10:05", "end": "11:20", "col": "E"},
    {"idx": 3, "label": "11:20-12:35", "start": "11:20", "end": "12:35", "col": "F"},
    {"idx": 4, "label": "13:10-14:25", "start": "13:10", "end": "14:25", "col": "H"},
    {"idx": 5, "label": "14:25-15:40", "start": "14:25", "end": "15:40", "col": "I"},
    {"idx": 6, "label": "15:40-16:55", "start": "15:40", "end": "16:55", "col": "J"},
    {"idx": 7, "label": "19:00-20:20", "start": "19:00", "end": "20:20", "col": "K"},
]


def _sql_str(v) -> str:
    if v is None:
        return "null"
    return "'" + str(v).strip().replace("'", "''") + "'"


def _sql_str_array(items) -> str:
    if not items:
        return "'{}'"
    inner = ",".join('"' + str(i).replace('"', '\\"') + '"' for i in items)
    return "'{" + inner + "}'"


def _clean_room(name: str) -> str | None:
    """Normalise a mined room token; return None for noise/placeholders."""
    s = str(name).strip()
    if not s or s == "***":
        return None
    s = s.replace("ACl", "ACL")  # lowercase-L typo
    return s


def mine_teachers() -> dict[str, dict]:
    """acronym -> {full_name, dept, designation, off_days[]}. Acronym set is
    authoritative from the distribution; names/off-days layered on top."""
    wb = openpyxl.load_workbook(DISTRIBUTION, data_only=True)
    ws = wb["Course Distribution"]
    teachers: dict[str, dict] = {}
    for r in range(6, ws.max_row + 1):
        t = ws.cell(row=r, column=9).value  # col I = Teacher
        if t is None or str(t).strip() == "":
            continue
        code = str(t).strip()
        teachers.setdefault(code, {"full_name": None, "dept": None,
                                   "designation": None, "off_days": []})

    wb2 = openpyxl.load_workbook(ROUTINE, data_only=True)

    # Names from Teacher Information (header row 3: C=Acronym, D=Full name,
    # E=Dept, F=Designation).
    ws = wb2["Teacher Information"]
    for r in range(4, ws.max_row + 1):
        ac = ws.cell(row=r, column=3).value
        if not ac:
            continue
        ac = str(ac).strip()
        if ac in teachers:
            teachers[ac]["full_name"] = (ws.cell(row=r, column=4).value or None)
            teachers[ac]["dept"] = (ws.cell(row=r, column=5).value or None)
            teachers[ac]["designation"] = (ws.cell(row=r, column=6).value or None)

    # GED / service names (header row 7: D=Acronym, E=Name, C=Department,
    # F=Designation).
    ws = wb2["GED Teacher Info"]
    for r in range(8, ws.max_row + 1):
        ac = ws.cell(row=r, column=4).value
        if not ac:
            continue
        ac = str(ac).strip()
        if ac in teachers and not teachers[ac]["full_name"]:
            teachers[ac]["full_name"] = (ws.cell(row=r, column=5).value or None)
            teachers[ac]["dept"] = (ws.cell(row=r, column=3).value or None)
            teachers[ac]["designation"] = (ws.cell(row=r, column=6).value or None)

    # Off-days from Teacher Day Off (header row 4: B=Teacher, C..I=Sun..Sat,
    # TRUE = OFF).
    ws = wb2["Teacher Day Off"]
    for r in range(5, ws.max_row + 1):
        ac = ws.cell(row=r, column=2).value
        if not ac:
            continue
        ac = str(ac).strip()
        if ac not in teachers:
            continue
        offs = []
        for i, day in enumerate(DAYS):
            if ws.cell(row=r, column=3 + i).value is True:
                offs.append(day)
        teachers[ac]["off_days"] = offs
    return teachers


def mine_rooms() -> list[dict]:
    """Room inventory: Room Acronym legend + every token mined from day cells."""
    wb = openpyxl.load_workbook(ROUTINE, data_only=True)
    found: set[str] = set()

    ws = wb["Room Acronym"]
    for r in range(1, ws.max_row + 1):
        c = _clean_room(ws.cell(row=r, column=2).value or "")
        if c and c not in ("Acronym",):
            found.add(c)

    code_re = re.compile(r"^\s*([A-Za-z]{2,4}-?\d{3,4})\s+(.*)$")
    for d in DAYS:
        ws = wb[d]
        for r in range(3, 62):
            for cc in PERIOD_COLS:
                v = ws.cell(row=r, column=cc).value
                if not v:
                    continue
                s = str(v).strip()
                if s.upper() in ("B", "R", "E", "A", "K", "BREAK"):
                    continue
                m = code_re.match(s)
                if m:
                    parts = m.group(2).split()
                    if len(parts) >= 2:
                        room = _clean_room(parts[-1])
                        if room:
                            found.add(room)

    rooms = []
    for name in sorted(found):
        is_lab = name in LAB_ROOMS
        is_gallery = name in GALLERIES
        building = None
        if name.startswith("RAB") or name in GALLERIES:
            building = "Ragib Ali Building"
        elif name.startswith("RKB") or name in ("GL",):
            building = "Rabeya Khatun Building"
        rooms.append({"name": name, "building": building,
                      "is_lab": is_lab, "is_gallery": is_gallery})
    return rooms


def main() -> None:
    teachers = mine_teachers()
    rooms = mine_rooms()

    settings = {
        "semester_label": "Summer 2025",
        "periods": PERIODS,
        "friday_no_p4": True,
        "service_scope": "resource_only",
        "weights": {"different_days": 8, "compactness": 3,
                    "spread": 2, "late_slot": 1},
    }

    # ── engine/config.json (engine fallback) ──
    config = {
        "days": DAYS,
        "periods": PERIODS,
        "friday_no_p4": True,
        "rooms": rooms,
        "lab_rooms": sorted(LAB_ROOMS),
        "galleries": sorted(GALLERIES),
        "teachers": teachers,
        "settings": settings,
    }
    (ENGINE_DIR / "config.json").write_text(
        json.dumps(config, indent=2, ensure_ascii=False), encoding="utf-8")

    # ── supabase/seed/seed_timetable_config.sql ──
    lines: list[str] = []
    lines.append("-- ============================================================")
    lines.append("-- SEED — timetable engine config (rooms, faculty, settings)")
    lines.append("-- Generated by engine/tools/seed_config.py from the real")
    lines.append("-- Summer-2025 distribution + routine workbooks. Idempotent.")
    lines.append("-- ============================================================\n")

    lines.append("-- Rooms")
    for r in rooms:
        lines.append(
            "insert into public.timetable_rooms (name, building, is_lab, is_gallery) values ("
            f"{_sql_str(r['name'])}, {_sql_str(r['building'])}, {str(r['is_lab']).lower()}, {str(r['is_gallery']).lower()})"
            " on conflict (name) do update set building=excluded.building, "
            "is_lab=excluded.is_lab, is_gallery=excluded.is_gallery;")

    lines.append("\n-- Faculty (76 acronyms; off_days TRUE=off per source sheet)")
    for ac in sorted(teachers):
        t = teachers[ac]
        lines.append(
            "insert into public.timetable_faculty (acronym, full_name, dept, designation, off_days) values ("
            f"{_sql_str(ac)}, {_sql_str(t['full_name'])}, {_sql_str(t['dept'])}, "
            f"{_sql_str(t['designation'])}, {_sql_str_array(t['off_days'])})"
            " on conflict (acronym) do update set full_name=excluded.full_name, "
            "dept=excluded.dept, designation=excluded.designation, off_days=excluded.off_days;")

    periods_json = json.dumps(settings["periods"]).replace("'", "''")
    weights_json = json.dumps(settings["weights"]).replace("'", "''")
    lines.append("\n-- Settings (single row)")
    lines.append(
        "insert into public.timetable_settings (id, semester_label, periods, friday_no_p4, service_scope, weights) values ("
        f"1, {_sql_str(settings['semester_label'])}, '{periods_json}'::jsonb, true, 'resource_only', '{weights_json}'::jsonb)"
        " on conflict (id) do update set semester_label=excluded.semester_label, "
        "periods=excluded.periods, friday_no_p4=excluded.friday_no_p4, "
        "service_scope=excluded.service_scope, weights=excluded.weights;")

    out = REPO_DIR / "supabase" / "seed" / "seed_timetable_config.sql"
    out.write_text("\n".join(lines) + "\n", encoding="utf-8")

    # ── summary ──
    print(f"teachers: {len(teachers)}  (with off-days: "
          f"{sum(1 for t in teachers.values() if t['off_days'])})")
    print(f"rooms: {len(rooms)}  (lab: {sum(1 for r in rooms if r['is_lab'])}, "
          f"gallery: {sum(1 for r in rooms if r['is_gallery'])})")
    print(f"wrote: {ENGINE_DIR / 'config.json'}")
    print(f"wrote: {out}")


if __name__ == "__main__":
    main()
